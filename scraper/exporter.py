import json
import os
import html
import urllib.parse
import argparse
import pandas as pd
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Colors for HTML Theme
PRIMARY_COLOR = "#0f172a"  # Slate 900
ACCENT_COLOR = "#2563eb"   # Blue 600
BG_COLOR = "#f8fafc"       # Slate 50
CARD_BG = "#ffffff"

# Localizations for HTML Pages
TRANSLATIONS = {
    "en": {
        "title": "Bogazici University MIS Personnel Directory",
        "subtitle": "Complete academic profiles and career details",
        "stat_total": "Total Personnel",
        "stat_ft": "Full-Time Faculty",
        "stat_pt": "Part-Time Faculty",
        "stat_contrib": "Contributing Faculty",
        "stat_ta": "TAs",
        "search_placeholder": "Search by name, title, or research interests...",
        "filter_all": "All",
        "filter_ft": "Full-Time",
        "filter_pt": "Part-Time",
        "filter_contrib": "Contributing",
        "filter_ta": "Teaching Assistants",
        "no_results_title": "No matching personnel found",
        "no_results_desc": "Try adjusting your search terms or filter category.",
        "edu": "Education",
        "ri": "Research Interests",
        "ct": "Courses Taught",
        "proj": "Projects",
        "source_url": "Source URL",
        "publications": "Publication",
        "publications_plural": "Publications",
        "back_btn": "Back to Directory",
        "cit_title": "Academic Publications & Citations",
        "cit_subtitle": "Department of Management Information Systems, Bogazici University",
        "cit_sidebar_search": "Filter by faculty name...",
        "cit_no_records": "No citation or publication records found for this faculty member on the MIS website.",
        "copy": "Copy",
        "copied": "Copied",
        "citations_html_name": "faculty_citations_en.html",
        "directory_html_name": "faculty_directory_en.html",
        "tbl_col_name": "Instructor",
        "tbl_col_role": "Role",
        "tbl_col_contact": "Contact",
        "tbl_col_source": "Source Profile",
        "tbl_col_publications": "Publications",
        "cv_bio": "CV / Biography",
        "cv_link_label": "Download CV (PDF)",
        "area": "Area",
        "recent_pub": "Recent Publications"
    },
    "tr": {
        "title": "Boğaziçi Üniversitesi YBS Akademik Kadro Dizini",
        "subtitle": "Detaylı akademik profiller ve kariyer bilgileri",
        "stat_total": "Toplam Kadro",
        "stat_ft": "Tam Zamanlı Kadro",
        "stat_pt": "Yarı Zamanlı Kadro",
        "stat_contrib": "Katkı Veren Akademisyenler",
        "stat_ta": "Araştırma Görevlileri",
        "search_placeholder": "İsim, unvan veya araştırma alanına göre ara...",
        "filter_all": "Tümü",
        "filter_ft": "Tam Zamanlı",
        "filter_pt": "Yarı Zamanlı",
        "filter_contrib": "Katkı Veren",
        "filter_ta": "Araştırma Görevlileri",
        "no_results_title": "Eşleşen akademik kadro bulunamadı",
        "no_results_desc": "Arama terimlerini veya filtre kategorisini değiştirmeyi deneyin.",
        "edu": "Eğitim",
        "ri": "Araştırma Alanları",
        "ct": "Verilen Dersler",
        "proj": "Projeler",
        "source_url": "Kaynak URL",
        "publications": "Yayın",
        "publications_plural": "Yayın",
        "back_btn": "Dizine Geri Dön",
        "cit_title": "Akademik Yayınlar & Atıflar",
        "cit_subtitle": "Yönetim Bilişim Sistemleri Bölümü, Boğaziçi Üniversitesi",
        "cit_sidebar_search": "Akademisyen ismine göre filtrele...",
        "cit_no_records": "Bu öğretim elemanı için MIS web sitesinde yayın veya atıf kaydı bulunamadı.",
        "copy": "Kopyala",
        "copied": "Kopyalandı",
        "citations_html_name": "faculty_citations_tr.html",
        "directory_html_name": "faculty_directory_tr.html",
        "tbl_col_name": "Öğretim Elemanı",
        "tbl_col_role": "Kategori",
        "tbl_col_contact": "İletişim",
        "tbl_col_source": "Kaynak Profil",
        "tbl_col_publications": "Yayınlar",
        "cv_bio": "Özgeçmiş / Biyografi",
        "cv_link_label": "CV İndir (PDF)",
        "area": "Alan",
        "recent_pub": "Son Yayınlar"
    }
}

# Excel Columns localization
EXCEL_COLS = {
    "en": {
        "sheet_dir": "Personnel Directory",
        "sheet_cit": "Citations Directory",
        "cols_dir": [
            "Name", "Title", "Role", "Email", "Phone", "Website", 
            "Education", "Courses Taught", "Research Interests", "Projects", 
            "CV Link", "Biography Text", "Area", "Recent Publications",
            "Profile URL", "Photo URL", "Total Citations"
        ],
        "cols_cit": ["Author", "Author Role", "Citation Category", "Citation Text"]
    },
    "tr": {
        "sheet_dir": "Akademik Kadro",
        "sheet_cit": "Yayınlar ve Atıflar",
        "cols_dir": [
            "Ad Soyad", "Unvan", "Kategori/Rol", "E-posta", "Telefon", "Web Sitesi", 
            "Eğitim", "Verilen Dersler", "Araştırma Alanları", "Projeler", 
            "Özgeçmiş Linki", "Biyografi Metni", "Alan", "Son Yayınlar",
            "Profil URL", "Fotoğraf URL", "Toplam Yayın Sayısı"
        ],
        "cols_cit": ["Akademisyen", "Akademisyen Rolü", "Yayın Kategorisi", "Yayın Metni"]
    }
}

def clean_html_text(text):
    if not text:
        return ""
    return html.escape(str(text))

def generate_excel(data, output_path, lang):
    print(f"[{lang.upper()}] Generating Excel directory to {output_path}...")
    
    cfg = EXCEL_COLS[lang]
    sheet_dir_name = cfg["sheet_dir"]
    sheet_cit_name = cfg["sheet_cit"]
    cols_dir = cfg["cols_dir"]
    cols_cit = cfg["cols_cit"]

    # 1. Directory Sheet Rows
    dir_rows = []
    for p in data:
        if "error" in p:
            continue
        
        cit_dict = p.get("citations", {})
        total_citations = sum(len(cits) for cits in cit_dict.values())
        
        row_data = [
            p.get("name", ""),
            p.get("title", ""),
            p.get("role", ""),
            p.get("email", ""),
            p.get("phone", ""),
            p.get("website", ""),
            p.get("education", ""),
            p.get("courses_taught", ""),
            p.get("research_interests", ""),
            p.get("projects", ""),
            p.get("cv_link", ""),
            p.get("cv_text", ""),
            p.get("area", ""),
            p.get("recent_publications", ""),
            p.get("url", ""),
            p.get("picture_url", ""),
            total_citations
        ]
        dir_rows.append(dict(zip(cols_dir, row_data)))
        
    df_dir = pd.DataFrame(dir_rows)

    # 2. Citations Sheet Rows
    cit_rows = []
    for p in data:
        if "error" in p or "citations" not in p:
            continue
        name = p.get("name", "")
        role = p.get("role", "")
        for cat_key, cits in p["citations"].items():
            cat_name = cat_key.replace("_", " ").title()
            if lang == "tr":
                cat_mapping = {
                    "international_articles": "Uluslararası Makale",
                    "international_book_chapters": "Uluslararası Kitap Bölümü",
                    "national_books": "Ulusal Kitap",
                    "national_articles": "Ulusal Makale",
                    "international_conference_papers": "Uluslararası Konferans Bildirisi",
                    "national_conference_papers": "Ulusal Konferans Bildirisi"
                }
                cat_name = cat_mapping.get(cat_key, cat_name)
                
            for cit in cits:
                row_data = [name, role, cat_name, cit]
                cit_rows.append(dict(zip(cols_cit, row_data)))
                
    df_cit = pd.DataFrame(cit_rows)

    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        df_dir.to_excel(writer, sheet_name=sheet_dir_name, index=False)
        df_cit.to_excel(writer, sheet_name=sheet_cit_name, index=False)

        workbook = writer.book
        ws_dir = writer.sheets[sheet_dir_name]
        ws_dir.views.sheetView[0].showGridLines = True
        
        header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
        thin_border = Border(
            left=Side(style='thin', color='D9D9D9'),
            right=Side(style='thin', color='D9D9D9'),
            top=Side(style='thin', color='D9D9D9'),
            bottom=Side(style='thin', color='D9D9D9')
        )
        
        for col_idx in range(1, ws_dir.max_column + 1):
            cell = ws_dir.cell(row=1, column=col_idx)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_align
            cell.border = thin_border
            
        ws_dir.row_dimensions[1].height = 28
        
        for r_idx in range(2, ws_dir.max_row + 1):
            ws_dir.row_dimensions[r_idx].height = 20
            for c_idx in range(1, ws_dir.max_column + 1):
                cell = ws_dir.cell(row=r_idx, column=c_idx)
                cell.border = thin_border
                col_name = ws_dir.cell(row=1, column=c_idx).value
                
                # Wrap long text fields
                if col_name in [cols_dir[6], cols_dir[7], cols_dir[8], cols_dir[9], cols_dir[11], cols_dir[12], cols_dir[13]]:
                    cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
                elif col_name in [cols_dir[0], cols_dir[1], cols_dir[2], cols_dir[3], cols_dir[4], cols_dir[5]]:
                    cell.alignment = Alignment(horizontal="left", vertical="center")
                else:
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                    
        for col in ws_dir.columns:
            col_letter = get_column_letter(col[0].column)
            col_name = col[0].value
            max_len = 0
            for cell in col:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
            
            if col_name in [cols_dir[6], cols_dir[7], cols_dir[8], cols_dir[9], cols_dir[11], cols_dir[13]]:
                ws_dir.column_dimensions[col_letter].width = 40
            elif col_name in [cols_dir[10], cols_dir[12], cols_dir[14], cols_dir[15]]:
                ws_dir.column_dimensions[col_letter].width = 25
            else:
                ws_dir.column_dimensions[col_letter].width = min(max(max_len + 3, 12), 30)

        ws_cit = writer.sheets[sheet_cit_name]
        ws_cit.views.sheetView[0].showGridLines = True
        
        for col_idx in range(1, ws_cit.max_column + 1):
            cell = ws_cit.cell(row=1, column=col_idx)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_align
            cell.border = thin_border
            
        ws_cit.row_dimensions[1].height = 28
        
        for r_idx in range(2, ws_cit.max_row + 1):
            ws_cit.row_dimensions[r_idx].height = 20
            for c_idx in range(1, ws_cit.max_column + 1):
                cell = ws_cit.cell(row=r_idx, column=c_idx)
                cell.border = thin_border
                col_name = ws_cit.cell(row=1, column=c_idx).value
                if col_name == cols_cit[3]:
                    cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
                else:
                    cell.alignment = Alignment(horizontal="left", vertical="center")
                    
        ws_cit.column_dimensions["A"].width = 25
        ws_cit.column_dimensions["B"].width = 20
        ws_cit.column_dimensions["C"].width = 25
        ws_cit.column_dimensions["D"].width = 80

    print(f"[{lang.upper()}] Excel file generated successfully.")

def generate_html_directory(data, output_path, lang):
    print(f"[{lang.upper()}] Generating HTML directory to {output_path}...")
    
    t = TRANSLATIONS[lang]
    
    # Calculate stats
    total_count = 0
    full_time = 0
    part_time = 0
    contrib = 0
    ta_count = 0
    
    faculty_list = []
    for p in data:
        if "error" in p:
            continue
        total_count += 1
        role = p.get("role", "")
        if "Full-Time" in role or "Tam Zamanlı" in role:
            full_time += 1
        if "Part-Time" in role or "Yarı Zamanlı" in role:
            part_time += 1
        if "Contributing" in role or "Katkı Veren" in role:
            contrib += 1
        if "Teaching Assistant" in role or "Araştırma Görevlileri" in role or "Araştırma Görevlisi" in role:
            ta_count += 1
            
        faculty_list.append(p)

    # Sort faculty by name
    faculty_list.sort(key=lambda x: x.get("name", ""))

    table_rows_html = []
    for p in faculty_list:
        name = clean_html_text(p.get("name", ""))
        title = clean_html_text(p.get("title", ""))
        if not title:
            title = "Araştırma Görevlisi" if lang == "tr" and ("Teaching Assistant" in p.get("role", "") or "Araştırma" in p.get("role", "")) else "No Title"
        role = clean_html_text(p.get("role", ""))
        email = clean_html_text(p.get("email", ""))
        phone = clean_html_text(p.get("phone", ""))
        website = p.get("website", "")
        profile_url = p.get("url", "")
        picture_url = p.get("picture_url", "")
        
        education = p.get("education", "")
        courses_taught = p.get("courses_taught", "")
        research_interests = p.get("research_interests", "")
        projects = p.get("projects", "")
        
        cv_link = p.get("cv_link", "")
        cv_text = p.get("cv_text", "")
        area = p.get("area", "")
        recent_publications = p.get("recent_publications", "")
        
        cit_dict = p.get("citations", {})
        cit_count = sum(len(cits) for cits in cit_dict.values())
        
        # Format text fields with clean inline linebreaks for direct table visibility
        def format_text_lines(text):
            if not text:
                return "<span class='no-data'>-</span>"
            return text.replace("\n", "<br>").strip()

        edu_formatted = format_text_lines(education)
        courses_formatted = format_text_lines(courses_taught)
        interests_formatted = format_text_lines(research_interests)
        projects_formatted = format_text_lines(projects)
        area_formatted = format_text_lines(area)
        recent_pub_formatted = format_text_lines(recent_publications)

        # CV / Bio column rendering
        cv_bio_parts = []
        if cv_link:
            cv_bio_parts.append(f'<div><a href="{clean_html_text(cv_link)}" target="_blank" rel="noopener" class="tbl-link"><i class="fa-solid fa-file-pdf"></i> {t["cv_link_label"]}</a></div>')
        if cv_text:
            # If text is long, render it, otherwise just display it
            if len(cv_text) > 40 or not cv_link:
                cv_bio_parts.append(f'<div class="tbl-cv-text-desc">{format_text_lines(cv_text)}</div>')
        cv_bio_formatted = "".join(cv_bio_parts) if cv_bio_parts else "<span class='no-data'>-</span>"

        # Avatar representation
        if picture_url:
            pic_html = f'<img class="tbl-profile-pic" src="{clean_html_text(picture_url)}" alt="{name}" loading="lazy">'
        else:
            initials = "".join([part[0] for part in name.split()[:2] if part]).upper()
            pic_html = f'<div class="tbl-profile-pic-placeholder">{initials}</div>'

        role_class = "role-ta"
        if "Full-Time" in role or "Tam Zamanlı" in role:
            role_class = "role-ft"
        elif "Part-Time" in role or "Yarı Zamanlı" in role:
            role_class = "role-pt"
        elif "Contributing" in role or "Katkı Veren" in role:
            role_class = "role-contrib"

        # Direct contacts html
        contacts = []
        if email:
            contacts.append(f'<div><a href="mailto:{email}" class="tbl-link"><i class="fa-regular fa-envelope"></i> {email}</a></div>')
        if phone:
            contacts.append(f'<div class="tbl-contact-text"><i class="fa-solid fa-phone"></i> {phone}</div>')
        if website:
            contacts.append(f'<div><a href="{clean_html_text(website)}" target="_blank" rel="noopener" class="tbl-link"><i class="fa-solid fa-globe"></i> Website</a></div>')
        contacts_html = "".join(contacts) if contacts else "<span class='no-data'>-</span>"

        enc_name = urllib.parse.quote_plus(p.get("name", ""))
        citation_label = t["publications_plural"] if cit_count != 1 else t["publications"]

        row_html = f"""
        <tr class="instructor-row" data-name="{name.lower()}" data-role="{role}" data-interests="{research_interests.lower()}" data-edu="{education.lower()}" data-courses="{courses_taught.lower()}" data-area="{area.lower()}" data-recent="{recent_publications.lower()}">
            <!-- 1. Instructor Name & Photo -->
            <td>
                <div class="instructor-meta">
                    <div class="tbl-pic-wrapper">
                        {pic_html}
                    </div>
                    <div>
                        <div class="instructor-name-field">{name}</div>
                        <div class="instructor-title-field">{title}</div>
                    </div>
                </div>
            </td>
            
            <!-- 2. Role -->
            <td style="white-space: nowrap;">
                <span class="role-badge {role_class}">{role}</span>
            </td>
            
            <!-- 3. Contact Info -->
            <td>
                <div class="tbl-contacts-cell">
                    {contacts_html}
                </div>
            </td>
            
            <!-- 4. Education -->
            <td class="tbl-text-cell">{edu_formatted}</td>
            
            <!-- 5. Research Interests -->
            <td class="tbl-text-cell">{interests_formatted}</td>
            
            <!-- 6. Courses Taught -->
            <td class="tbl-text-cell">{courses_formatted}</td>
            
            <!-- 7. Projects -->
            <td class="tbl-text-cell">{projects_formatted}</td>
            
            <!-- 8. CV / Biography -->
            <td class="tbl-text-cell">{cv_bio_formatted}</td>
            
            <!-- 9. Area -->
            <td class="tbl-text-cell">{area_formatted}</td>
            
            <!-- 10. Recent Publications -->
            <td class="tbl-text-cell">{recent_pub_formatted}</td>
            
            <!-- 11. Publications Link -->
            <td style="white-space: nowrap; text-align: center; vertical-align: middle;">
                <a href="{t["citations_html_name"]}?author={enc_name}" class="tbl-citations-btn">
                    <i class="fa-solid fa-quote-right"></i> {cit_count} {citation_label}
                </a>
            </td>

            <!-- 12. Source Profile Link (Scraped URL) -->
            <td>
                <a href="{clean_html_text(profile_url)}" target="_blank" rel="noopener" class="tbl-source-link" title="{clean_html_text(profile_url)}">
                    <i class="fa-solid fa-arrow-up-right-from-square"></i> Link
                </a>
            </td>
        </tr>
        """
        table_rows_html.append(row_html)

    all_rows_html = "\n".join(table_rows_html)

    html_content = f"""<!DOCTYPE html>
<html lang="{lang}">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>{t["title"]}</title>
    <link rel="stylesheet" href="https://cdnjs.cloudflare.com/ajax/libs/font-awesome/6.4.0/css/all.min.css">
    <link rel="preconnect" href="https://fonts.googleapis.com">
    <link rel="preconnect" href="https://fonts.gstatic.com" crossorigin>
    <link href="https://fonts.googleapis.com/css2?family=Inter:wght@300;400;500;600;700&display=swap" rel="stylesheet">
    
    <style>
        :root {{
            --bg-color: {BG_COLOR};
            --primary-color: {PRIMARY_COLOR};
            --accent-color: {ACCENT_COLOR};
            --card-bg: {CARD_BG};
            --text-main: #334155;
            --text-muted: #64748b;
            --border-color: #e2e8f0;
            --radius-lg: 16px;
            --radius-md: 8px;
            --shadow-sm: 0 1px 3px 0 rgba(0, 0, 0, 0.1), 0 1px 2px -1px rgba(0, 0, 0, 0.1);
            --shadow-md: 0 4px 6px -1px rgba(0, 0, 0, 0.1), 0 2px 4px -2px rgba(0, 0, 0, 0.1);
        }}

        * {{
            box-sizing: border-box;
            margin: 0;
            padding: 0;
        }}

        body {{
            font-family: 'Inter', sans-serif;
            background-color: var(--bg-color);
            color: var(--text-main);
            line-height: 1.5;
            padding: 2rem 1rem;
        }}

        .container {{
            max-width: 1750px;
            margin: 0 auto;
        }}

        header {{
            margin-bottom: 2rem;
            text-align: center;
        }}

        header h1 {{
            font-size: 2rem;
            font-weight: 700;
            color: var(--primary-color);
            margin-bottom: 0.5rem;
            letter-spacing: -0.025em;
        }}

        header p {{
            color: var(--text-muted);
            font-size: 1rem;
        }}

        .stats-dashboard {{
            display: grid;
            grid-template-columns: repeat(auto-fit, minmax(150px, 1fr));
            gap: 1rem;
            margin-bottom: 2rem;
        }}

        .stat-card {{
            background: var(--card-bg);
            border: 1px solid var(--border-color);
            border-radius: var(--radius-md);
            padding: 1rem;
            text-align: center;
            box-shadow: var(--shadow-sm);
        }}

        .stat-num {{
            font-size: 1.75rem;
            font-weight: 700;
            color: var(--accent-color);
        }}

        .stat-label {{
            font-size: 0.75rem;
            font-weight: 600;
            color: var(--text-muted);
            text-transform: uppercase;
            letter-spacing: 0.05em;
        }}

        .controls-panel {{
            background: var(--card-bg);
            border: 1px solid var(--border-color);
            border-radius: var(--radius-md);
            padding: 1.25rem;
            margin-bottom: 2rem;
            box-shadow: var(--shadow-sm);
            display: flex;
            flex-direction: column;
            gap: 1rem;
        }}

        @media (min-width: 768px) {{
            .controls-panel {{
                flex-direction: row;
                align-items: center;
                justify-content: space-between;
            }}
        }}

        .search-wrapper {{
            position: relative;
            flex-grow: 1;
            max-width: 500px;
        }}

        .search-wrapper i {{
            position: absolute;
            left: 1rem;
            top: 50%;
            transform: translateY(-50%);
            color: var(--text-muted);
        }}

        .search-input {{
            width: 100%;
            padding: 0.65rem 1rem 0.65rem 2.5rem;
            border: 1px solid var(--border-color);
            border-radius: var(--radius-md);
            font-size: 0.9rem;
            font-family: inherit;
            color: inherit;
            background-color: var(--bg-color);
        }}

        .search-input:focus {{
            outline: none;
            border-color: var(--accent-color);
            background-color: #fff;
            box-shadow: 0 0 0 3px rgba(37, 99, 235, 0.15);
        }}

        .filter-group {{
            display: flex;
            flex-wrap: wrap;
            gap: 0.5rem;
        }}

        .filter-btn {{
            background-color: var(--bg-color);
            border: 1px solid var(--border-color);
            padding: 0.4rem 0.85rem;
            border-radius: 9999px;
            font-size: 0.8rem;
            font-weight: 500;
            color: var(--text-muted);
            cursor: pointer;
            transition: all 0.2s;
        }}

        .filter-btn:hover {{
            background-color: #f1f5f9;
            color: var(--text-main);
        }}

        .filter-btn.active {{
            background-color: var(--accent-color);
            color: white;
            border-color: var(--accent-color);
        }}

        /* Table Design */
        .table-responsive {{
            width: 100%;
            overflow-x: auto;
            background: var(--card-bg);
            border: 1px solid var(--border-color);
            border-radius: var(--radius-md);
            box-shadow: var(--shadow-sm);
            margin-bottom: 2rem;
        }}

        table {{
            width: 100%;
            border-collapse: collapse;
            text-align: left;
            font-size: 0.825rem;
        }}

        th {{
            background-color: var(--primary-color);
            color: white;
            padding: 0.85rem 1rem;
            font-weight: 600;
            font-size: 0.825rem;
            border: none;
            white-space: nowrap;
        }}

        td {{
            padding: 0.85rem;
            border-bottom: 1px solid var(--border-color);
            vertical-align: top;
            line-height: 1.5;
        }}

        tr:hover {{
            background-color: #f8fafc;
        }}

        /* Photo and Meta in cell */
        .instructor-meta {{
            display: flex;
            align-items: center;
            gap: 0.85rem;
            min-width: 200px;
        }}

        .tbl-pic-wrapper {{
            width: 48px;
            height: 48px;
            border-radius: 50%;
            overflow: hidden;
            flex-shrink: 0;
            border: 1px solid var(--border-color);
            background-color: var(--bg-color);
            display: flex;
            align-items: center;
            justify-content: center;
        }}

        .tbl-profile-pic {{
            width: 100%;
            height: 100%;
            object-fit: cover;
        }}

        .tbl-profile-pic-placeholder {{
            font-size: 1.1rem;
            font-weight: 700;
            color: var(--accent-color);
            background: rgba(37, 99, 235, 0.1);
            width: 100%;
            height: 100%;
            display: flex;
            align-items: center;
            justify-content: center;
        }}

        .instructor-name-field {{
            font-weight: 700;
            color: var(--primary-color);
            font-size: 0.95rem;
        }}

        .instructor-title-field {{
            font-size: 0.75rem;
            color: var(--text-muted);
            margin-top: 0.1rem;
        }}

        .role-badge {{
            display: inline-block;
            padding: 0.15rem 0.5rem;
            border-radius: 9999px;
            font-size: 0.7rem;
            font-weight: 600;
            text-transform: uppercase;
            letter-spacing: 0.025em;
        }}

        .role-ft {{ background-color: #dbeafe; color: #1e40af; }}
        .role-pt {{ background-color: #f3e8ff; color: #6b21a8; }}
        .role-contrib {{ background-color: #ccfbf1; color: #115e59; }}
        .role-ta {{ background-color: #ffedd5; color: #9a3412; }}

        .tbl-contacts-cell {{
            display: flex;
            flex-direction: column;
            gap: 0.35rem;
            min-width: 165px;
        }}

        .tbl-link {{
            color: var(--accent-color);
            text-decoration: none;
            display: inline-flex;
            align-items: center;
            gap: 0.35rem;
        }}

        .tbl-link:hover {{
            text-decoration: underline;
        }}

        .tbl-contact-text {{
            color: var(--text-main);
            display: inline-flex;
            align-items: center;
            gap: 0.35rem;
        }}

        .tbl-text-cell {{
            min-width: 180px;
            max-width: 320px;
            font-size: 0.8rem;
            color: var(--text-main);
            word-wrap: break-word;
        }}
        
        .tbl-cv-text-desc {{
            margin-top: 0.35rem;
            max-height: 120px;
            overflow-y: auto;
            font-size: 0.75rem;
            color: var(--text-muted);
            border: 1px solid var(--border-color);
            padding: 0.35rem;
            border-radius: var(--radius-md);
            background-color: var(--bg-color);
        }}

        .no-data {{
            color: var(--text-muted);
            font-style: italic;
        }}

        .tbl-citations-btn {{
            display: inline-flex;
            align-items: center;
            gap: 0.35rem;
            background-color: var(--primary-color);
            color: white;
            text-decoration: none;
            padding: 0.4rem 0.75rem;
            border-radius: var(--radius-md);
            font-size: 0.75rem;
            font-weight: 600;
            white-space: nowrap;
        }}

        .tbl-citations-btn:hover {{
            opacity: 0.9;
        }}

        .tbl-source-link {{
            color: var(--text-muted);
            text-decoration: none;
            display: inline-flex;
            align-items: center;
            gap: 0.25rem;
        }}

        .tbl-source-link:hover {{
            color: var(--accent-color);
            text-decoration: underline;
        }}

        .no-results {{
            text-align: center;
            padding: 4rem 2rem;
            color: var(--text-muted);
            background: white;
            display: none;
        }}

        .no-results i {{
            font-size: 2.5rem;
            margin-bottom: 1rem;
            color: #cbd5e1;
        }}
        
        .lang-switch {{
            margin-bottom: 1.25rem;
            display: flex;
            justify-content: center;
            gap: 1rem;
        }}
        
        .lang-link {{
            color: var(--text-muted);
            text-decoration: none;
            font-size: 0.85rem;
            font-weight: 600;
        }}
        
        .lang-link.active {{
            color: var(--accent-color);
            border-bottom: 2px solid var(--accent-color);
        }}
    </style>
</head>
<body>
    <div class="container">
        <!-- Language Switcher -->
        <div class="lang-switch">
            <a href="faculty_directory_en.html" class="lang-link {"active" if lang == "en" else ""}">English</a>
            <a href="faculty_directory_tr.html" class="lang-link {"active" if lang == "tr" else ""}">Türkçe</a>
        </div>

        <header>
            <h1>{t["title"]}</h1>
            <p>{t["subtitle"]}</p>
        </header>

        <!-- Stats Dashboard -->
        <section class="stats-dashboard">
            <div class="stat-card">
                <div class="stat-num">{total_count}</div>
                <div class="stat-label">{t["stat_total"]}</div>
            </div>
            <div class="stat-card">
                <div class="stat-num">{full_time}</div>
                <div class="stat-label">{t["stat_ft"]}</div>
            </div>
            <div class="stat-card">
                <div class="stat-num">{part_time}</div>
                <div class="stat-label">{t["stat_pt"]}</div>
            </div>
            <div class="stat-card">
                <div class="stat-num">{contrib}</div>
                <div class="stat-label">{t["stat_contrib"]}</div>
            </div>
            <div class="stat-card">
                <div class="stat-num">{ta_count}</div>
                <div class="stat-label">{t["stat_ta"]}</div>
            </div>
        </section>

        <!-- Controls -->
        <section class="controls-panel">
            <div class="search-wrapper">
                <i class="fa-solid fa-magnifying-glass"></i>
                <input type="text" id="searchInput" class="search-input" placeholder="{t["search_placeholder"]}">
            </div>
            
            <div class="filter-group">
                <button class="filter-btn active" onclick="filterRole('All')">{t["filter_all"]}</button>
                <button class="filter-btn" onclick="filterRole('Faculty')">{t["filter_ft"]}</button>
                <button class="filter-btn" onclick="filterRole('Yarı')">{t["filter_pt"]}</button>
                <button class="filter-btn" onclick="filterRole('Katkı')">{t["filter_contrib"]}</button>
                <button class="filter-btn" onclick="filterRole('Görevli')">{t["filter_ta"]}</button>
            </div>
        </section>

        <!-- Plain Data Table -->
        <div class="table-responsive">
            <table id="instructorTable">
                <thead>
                    <tr>
                        <th>{t["tbl_col_name"]}</th>
                        <th>{t["tbl_col_role"]}</th>
                        <th>{t["tbl_col_contact"]}</th>
                        <th>{t["edu"]}</th>
                        <th>{t["ri"]}</th>
                        <th>{t["ct"]}</th>
                        <th>{t["proj"]}</th>
                        <th>{t["cv_bio"]}</th>
                        <th>{t["area"]}</th>
                        <th>{t["recent_pub"]}</th>
                        <th>{t["tbl_col_publications"]}</th>
                        <th>{t["tbl_col_source"]}</th>
                    </tr>
                </thead>
                <tbody>
                    {all_rows_html}
                </tbody>
            </table>
            
            <div class="no-results" id="noResults">
                <i class="fa-solid fa-user-slash"></i>
                <h3>{t["no_results_title"]}</h3>
                <p>{t["no_results_desc"]}</p>
            </div>
        </div>
    </div>

    <script>
        let currentRole = 'All';
        
        function filterRole(role) {{
            currentRole = role;
            
            const buttons = document.querySelectorAll('.filter-btn');
            buttons.forEach(btn => {{
                if (btn.innerText.trim() === '{t["filter_all"]}' && role === 'All') {{
                    btn.classList.add('active');
                }} else if (btn.innerText.trim() === '{t["filter_ft"]}' && role === 'Faculty') {{
                    btn.classList.add('active');
                }} else if (btn.innerText.trim() === '{t["filter_pt"]}' && role === 'Yarı') {{
                    btn.classList.add('active');
                }} else if (btn.innerText.trim() === '{t["filter_contrib"]}' && role === 'Katkı') {{
                    btn.classList.add('active');
                }} else if (btn.innerText.trim() === '{t["filter_ta"]}' && role === 'Görevli') {{
                    btn.classList.add('active');
                }} else {{
                    btn.classList.remove('active');
                }}
            }});
            
            applySearchAndFilter();
        }}

        const searchInput = document.getElementById('searchInput');
        searchInput.addEventListener('input', applySearchAndFilter);

        function applySearchAndFilter() {{
            const query = searchInput.value.toLowerCase().trim();
            const rows = document.querySelectorAll('.instructor-row');
            let visibleCount = 0;
            
            rows.forEach(row => {{
                const name = row.getAttribute('data-name');
                const rowRole = row.getAttribute('data-role');
                const interests = row.getAttribute('data-interests');
                const edu = row.getAttribute('data-edu');
                const courses = row.getAttribute('data-courses');
                const area = row.getAttribute('data-area') || '';
                const recent = row.getAttribute('data-recent') || '';
                
                // Check Role Match
                let roleMatch = false;
                if (currentRole === 'All') {{
                    roleMatch = true;
                }} else if (currentRole === 'Faculty') {{
                    roleMatch = rowRole.includes('Faculty') || rowRole.includes('Kadro');
                }} else if (currentRole === 'Yarı') {{
                    roleMatch = rowRole.includes('Part-Time') || rowRole.includes('Yarı Zamanlı');
                }} else if (currentRole === 'Katkı') {{
                    roleMatch = rowRole.includes('Contributing') || rowRole.includes('Katkı');
                }} else if (currentRole === 'Görevli') {{
                    roleMatch = rowRole.includes('Assistant') || rowRole.includes('Görevli') || rowRole.includes('Görevlileri');
                }}
                
                // Check Search Match
                let searchMatch = false;
                if (query === '') {{
                    searchMatch = true;
                }} else {{
                    searchMatch = name.includes(query) || 
                                  interests.includes(query) || 
                                  rowRole.toLowerCase().includes(query) ||
                                  edu.includes(query) ||
                                  courses.includes(query) ||
                                  area.includes(query) ||
                                  recent.includes(query);
                }}
                
                if (roleMatch && searchMatch) {{
                    row.style.display = 'table-row';
                    visibleCount++;
                }} else {{
                    row.style.display = 'none';
                }}
            }});
            
            const table = document.getElementById('instructorTable');
            const noResults = document.getElementById('noResults');
            
            if (visibleCount === 0) {{
                table.style.display = 'none';
                noResults.style.display = 'block';
            }} else {{
                table.style.display = 'table';
                noResults.style.display = 'none';
            }}
        }}
    </script>
</body>
</html>
"""
    with open(output_path, "w", encoding="utf-8") as f:
        f.write(html_content)
    print(f"[{lang.upper()}] HTML Directory generated successfully.")

def generate_html_citations(data, output_path, lang):
    print(f"[{lang.upper()}] Generating HTML citations to {output_path}...")
    
    t = TRANSLATIONS[lang]
    sorted_faculty = sorted([p for p in data if "error" not in p], key=lambda x: x.get("name", ""))
    
    total_cits = 0
    sidebar_items = []
    main_citations_html = []
    
    for idx, p in enumerate(sorted_faculty):
        name = clean_html_text(p.get("name", ""))
        role = clean_html_text(p.get("role", ""))
        profile_url = p.get("url", "")
        
        cit_dict = p.get("citations", {})
        cit_count = sum(len(cits) for cits in cit_dict.values())
        total_cits += cit_count
        
        active_class = "active" if idx == 0 else ""
        display_style = "block" if idx == 0 else "none"
        
        sidebar_items.append(f"""
        <button class="author-item {active_class}" onclick="selectAuthor('{clean_html_text(p.get("name", ""))}', this)" data-name="{name.lower()}">
            <span class="author-name">{name}</span>
            <span class="author-badge">{cit_count}</span>
        </button>
        """)

        details_html = []
        
        categories_map = {
            "international_articles": "International Articles" if lang == "en" else "Uluslararası Makaleler",
            "international_book_chapters": "International Book Chapters" if lang == "en" else "Uluslararası Kitap Bölümleri",
            "national_books": "National Books" if lang == "en" else "Ulusal Kitaplar",
            "national_articles": "National Articles" if lang == "en" else "Ulusal Makaleler",
            "international_conference_papers": "International Conference Papers" if lang == "en" else "Uluslararası Bildiriler",
            "national_conference_papers": "National Conference Papers" if lang == "en" else "Ulusal Bildiriler"
        }
        
        has_any_citations = False
        for cat_key, cat_name in categories_map.items():
            citations = cit_dict.get(cat_key, [])
            if citations:
                has_any_citations = True
                cit_items_html = []
                for cit in citations:
                    cit_cleaned = clean_html_text(cit)
                    cit_items_html.append(f"""
                    <div class="citation-entry">
                        <div class="cit-text">{cit_cleaned}</div>
                        <button class="copy-cit-btn" onclick="copyCitation(this, `{html.escape(cit)}`)">
                            <i class="fa-regular fa-copy"></i> {t["copy"]}
                        </button>
                    </div>
                    """)
                
                details_html.append(f"""
                <div class="citation-category-section">
                    <h4 class="category-title"><i class="fa-solid fa-tag"></i> {cat_name} <span class="category-count">({len(citations)})</span></h4>
                    <div class="citation-list">
                        {"".join(cit_items_html)}
                    </div>
                </div>
                """)
                
        if not has_any_citations:
            details_html.append(f"""
            <div class="no-citations-message">
                <i class="fa-solid fa-quote-left" style="font-size: 2.5rem; color: #cbd5e1; margin-bottom: 0.75rem;"></i>
                <p>{t["cit_no_records"]}</p>
            </div>
            """)

        main_citations_html.append(f"""
        <div class="author-citations-panel" id="citations-panel-{clean_html_text(p.get("name", ""))}" style="display: {display_style};">
            <div class="author-details-header">
                <div>
                    <h2 class="author-panel-name">{name}</h2>
                    <div class="author-panel-role">{role}</div>
                </div>
                <div class="author-panel-stats">
                    <div class="stat-box">
                        <div class="stat-box-val">{cit_count}</div>
                        <div class="stat-box-lbl">{t["publications"]}</div>
                    </div>
                    <a href="{clean_html_text(profile_url)}" target="_blank" class="view-profile-link-btn" title="View Profile">
                        <i class="fa-solid fa-arrow-up-right-from-square"></i> Profile
                    </a>
                </div>
            </div>
            
            <div class="panel-scraped-url">
                <span class="url-label">{t["source_url"]}:</span>
                <a href="{clean_html_text(profile_url)}" target="_blank" class="url-link">{clean_html_text(profile_url)}</a>
            </div>
            
            <div class="citations-container">
                {"".join(details_html)}
            </div>
        </div>
        """)

    all_sidebar_html = "\n".join(sidebar_items)
    all_main_html = "\n".join(main_citations_html)

    html_content = f"""<!DOCTYPE html>
<html lang="{lang}">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>{t["cit_title"]}</title>
    <link rel="stylesheet" href="https://cdnjs.cloudflare.com/ajax/libs/font-awesome/6.4.0/css/all.min.css">
    <link rel="preconnect" href="https://fonts.googleapis.com">
    <link rel="preconnect" href="https://fonts.gstatic.com" crossorigin>
    <link href="https://fonts.googleapis.com/css2?family=Inter:wght@300;400;500;600;700&display=swap" rel="stylesheet">
    
    <style>
        :root {{
            --bg-color: {BG_COLOR};
            --primary-color: {PRIMARY_COLOR};
            --accent-color: {ACCENT_COLOR};
            --card-bg: {CARD_BG};
            --text-main: #334155;
            --text-muted: #64748b;
            --border-color: #e2e8f0;
            --radius-lg: 16px;
            --radius-md: 8px;
            --shadow-sm: 0 1px 3px 0 rgba(0, 0, 0, 0.1), 0 1px 2px -1px rgba(0, 0, 0, 0.1);
            --shadow-md: 0 4px 6px -1px rgba(0, 0, 0, 0.1), 0 2px 4px -2px rgba(0, 0, 0, 0.1);
        }}

        * {{
            box-sizing: border-box;
            margin: 0;
            padding: 0;
        }}

        body {{
            font-family: 'Inter', sans-serif;
            background-color: var(--bg-color);
            color: var(--text-main);
            line-height: 1.5;
            height: 100vh;
            display: flex;
            flex-direction: column;
            overflow: hidden;
        }}

        .app-header {{
            background-color: var(--primary-color);
            color: white;
            padding: 1rem 2rem;
            display: flex;
            align-items: center;
            justify-content: space-between;
            box-shadow: var(--shadow-md);
            flex-shrink: 0;
        }}

        .logo-section h1 {{
            font-size: 1.35rem;
            font-weight: 700;
            letter-spacing: -0.025em;
            display: flex;
            align-items: center;
            gap: 0.5rem;
        }}

        .logo-section p {{
            color: #94a3b8;
            font-size: 0.8rem;
        }}

        .back-btn {{
            background: rgba(255, 255, 255, 0.1);
            color: white;
            text-decoration: none;
            padding: 0.5rem 1rem;
            border-radius: var(--radius-md);
            font-size: 0.875rem;
            font-weight: 600;
            display: flex;
            align-items: center;
            gap: 0.5rem;
            transition: background 0.2s;
        }}

        .back-btn:hover {{
            background: rgba(255, 255, 255, 0.2);
        }}

        .app-layout {{
            display: flex;
            flex-grow: 1;
            overflow: hidden;
        }}

        .app-sidebar {{
            width: 320px;
            background: white;
            border-right: 1px solid var(--border-color);
            display: flex;
            flex-direction: column;
            flex-shrink: 0;
        }}

        .sidebar-search {{
            padding: 1rem;
            border-bottom: 1px solid var(--border-color);
            position: relative;
        }}

        .sidebar-search i {{
            position: absolute;
            left: 1.75rem;
            top: 50%;
            transform: translateY(-50%);
            color: var(--text-muted);
        }}

        .sidebar-search input {{
            width: 100%;
            padding: 0.65rem 1rem 0.65rem 2.5rem;
            border: 1px solid var(--border-color);
            border-radius: var(--radius-md);
            font-size: 0.875rem;
            font-family: inherit;
            background-color: var(--bg-color);
        }}

        .sidebar-search input:focus {{
            outline: none;
            border-color: var(--accent-color);
            background-color: white;
        }}

        .author-list {{
            flex-grow: 1;
            overflow-y: auto;
            padding: 0.5rem;
            display: flex;
            flex-direction: column;
            gap: 0.25rem;
        }}

        .author-item {{
            background: none;
            border: none;
            padding: 0.75rem 1rem;
            border-radius: var(--radius-md);
            text-align: left;
            cursor: pointer;
            display: flex;
            align-items: center;
            justify-content: space-between;
            transition: all 0.15s;
            font-family: inherit;
            color: var(--text-main);
        }}

        .author-item:hover {{
            background-color: var(--bg-color);
        }}

        .author-item.active {{
            background-color: rgba(37, 99, 235, 0.08);
            color: var(--accent-color);
            font-weight: 600;
        }}

        .author-name {{
            font-size: 0.9rem;
            white-space: nowrap;
            overflow: hidden;
            text-overflow: ellipsis;
            padding-right: 0.5rem;
        }}

        .author-badge {{
            background-color: #cbd5e1;
            color: #475569;
            font-size: 0.75rem;
            font-weight: 700;
            padding: 0.15rem 0.5rem;
            border-radius: 9999px;
            transition: all 0.15s;
        }}

        .author-item.active .author-badge {{
            background-color: var(--accent-color);
            color: white;
        }}

        .app-content {{
            flex-grow: 1;
            overflow-y: auto;
            background-color: var(--bg-color);
            padding: 2rem;
        }}

        .author-citations-panel {{
            max-width: 900px;
            margin: 0 auto;
        }}

        .author-details-header {{
            background-color: white;
            border: 1px solid var(--border-color);
            border-radius: var(--radius-lg);
            padding: 1.5rem;
            box-shadow: var(--shadow-sm);
            display: flex;
            flex-direction: column;
            gap: 1.25rem;
            margin-bottom: 1.5rem;
        }}

        @media (min-width: 600px) {{
            .author-details-header {{
                flex-direction: row;
                align-items: center;
                justify-content: space-between;
            }}
        }}

        .author-panel-name {{
            font-size: 1.75rem;
            font-weight: 700;
            color: var(--primary-color);
            letter-spacing: -0.02em;
        }}

        .author-panel-role {{
            font-size: 0.95rem;
            color: var(--text-muted);
            font-weight: 500;
            margin-top: 0.15rem;
        }}

        .author-panel-stats {{
            display: flex;
            align-items: center;
            gap: 1rem;
        }}

        .stat-box {{
            text-align: center;
            background-color: var(--bg-color);
            border: 1px solid var(--border-color);
            border-radius: var(--radius-md);
            padding: 0.5rem 1rem;
            min-width: 100px;
        }}

        .stat-box-val {{
            font-size: 1.5rem;
            font-weight: 700;
            color: var(--accent-color);
        }}

        .stat-box-lbl {{
            font-size: 0.75rem;
            font-weight: 600;
            color: var(--text-muted);
            text-transform: uppercase;
        }}

        .view-profile-link-btn {{
            background-color: #10b981;
            color: white;
            text-decoration: none;
            padding: 0.65rem 1.25rem;
            border-radius: var(--radius-md);
            font-size: 0.875rem;
            font-weight: 600;
            display: flex;
            align-items: center;
            gap: 0.5rem;
            transition: opacity 0.2s;
        }}

        .view-profile-link-btn:hover {{
            opacity: 0.9;
        }}

        .panel-scraped-url {{
            font-size: 0.8rem;
            background-color: #f1f5f9;
            border: 1px solid var(--border-color);
            border-radius: var(--radius-md);
            padding: 0.65rem 1rem;
            margin-bottom: 2rem;
            display: flex;
            align-items: center;
            overflow: hidden;
            text-overflow: ellipsis;
            white-space: nowrap;
        }}

        .panel-scraped-url .url-label {{
            font-weight: 600;
            color: var(--text-muted);
            margin-right: 0.5rem;
        }}

        .panel-scraped-url .url-link {{
            color: var(--accent-color);
            text-decoration: none;
        }}

        .panel-scraped-url .url-link:hover {{
            text-decoration: underline;
        }}

        .citation-category-section {{
            margin-bottom: 2rem;
        }}

        .category-title {{
            font-size: 1.1rem;
            font-weight: 700;
            color: var(--primary-color);
            margin-bottom: 1rem;
            display: flex;
            align-items: center;
            gap: 0.5rem;
        }}

        .category-count {{
            font-size: 0.875rem;
            color: var(--text-muted);
            font-weight: 500;
        }}

        .citation-list {{
            display: flex;
            flex-direction: column;
            gap: 1rem;
        }}

        .citation-entry {{
            background-color: white;
            border: 1px solid var(--border-color);
            border-radius: var(--radius-md);
            padding: 1.25rem;
            box-shadow: var(--shadow-sm);
            position: relative;
            display: flex;
            justify-content: space-between;
            align-items: flex-start;
            gap: 1.5rem;
            transition: box-shadow 0.2s;
        }}

        .citation-entry:hover {{
            box-shadow: var(--shadow-md);
        }}

        .cit-text {{
            font-size: 0.925rem;
            color: var(--text-main);
            line-height: 1.6;
        }}

        .copy-cit-btn {{
            background-color: var(--bg-color);
            border: 1px solid var(--border-color);
            color: var(--text-muted);
            padding: 0.4rem 0.8rem;
            border-radius: var(--radius-md);
            font-size: 0.75rem;
            font-weight: 600;
            cursor: pointer;
            display: flex;
            align-items: center;
            gap: 0.35rem;
            transition: all 0.2s;
            flex-shrink: 0;
        }}

        .copy-cit-btn:hover {{
            background-color: var(--accent-color);
            color: white;
            border-color: var(--accent-color);
        }}

        .no-citations-message {{
            background-color: white;
            border: 1px solid var(--border-color);
            border-radius: var(--radius-lg);
            padding: 4rem 2rem;
            text-align: center;
            color: var(--text-muted);
        }}

        .toast {{
            position: fixed;
            bottom: 2rem;
            left: 50%;
            transform: translateX(-50%) translateY(100px);
            background-color: #1e293b;
            color: white;
            padding: 0.75rem 1.5rem;
            border-radius: 9999px;
            font-size: 0.875rem;
            font-weight: 500;
            box-shadow: var(--shadow-lg);
            display: flex;
            align-items: center;
            gap: 0.5rem;
            opacity: 0;
            transition: transform 0.3s cubic-bezier(0.16, 1, 0.3, 1), opacity 0.3s;
            z-index: 1000;
        }}

        .toast.show {{
            transform: translateX(-50%) translateY(0);
            opacity: 1;
        }}
    </style>
</head>
<body>
    <div class="app-header">
        <div class="logo-section">
            <h1><i class="fa-solid fa-quote-left" style="color: #60a5fa;"></i> {t["cit_title"]}</h1>
            <p>{t["cit_subtitle"]}</p>
        </div>
        <a href="{t["directory_html_name"]}" class="back-btn"><i class="fa-solid fa-chevron-left"></i> {t["back_btn"]}</a>
    </div>

    <div class="app-layout">
        <!-- Sidebar -->
        <div class="app-sidebar">
            <div class="sidebar-search">
                <i class="fa-solid fa-magnifying-glass"></i>
                <input type="text" id="sidebarSearchInput" placeholder="{t["cit_sidebar_search"]}">
            </div>
            
            <div class="author-list" id="authorList">
                {all_sidebar_html}
            </div>
        </div>
        
        <!-- Content Panel -->
        <div class="app-content" id="appContent">
            {all_main_html}
        </div>
    </div>

    <div class="toast" id="copyToast">
        <i class="fa-solid fa-circle-check" style="color: #34d399;"></i> {t["copied"]}!
    </div>

    <script>
        let currentAuthorName = "";
        
        const firstActive = document.querySelector('.author-item.active');
        if (firstActive) {{
            const nameSpan = firstActive.querySelector('.author-name');
            if (nameSpan) {{
                currentAuthorName = nameSpan.innerText.trim();
            }}
        }}

        function selectAuthor(authorName, btnElement) {{
            if (currentAuthorName === authorName) return;
            
            const prevPanel = document.getElementById('citations-panel-' + currentAuthorName);
            if (prevPanel) prevPanel.style.display = 'none';
            
            const newPanel = document.getElementById('citations-panel-' + authorName);
            if (newPanel) newPanel.style.display = 'block';
            
            const prevActiveBtn = document.querySelector('.author-item.active');
            if (prevActiveBtn) prevActiveBtn.classList.remove('active');
            
            btnElement.classList.add('active');
            currentAuthorName = authorName;
            document.getElementById('appContent').scrollTop = 0;
        }}

        const sidebarSearchInput = document.getElementById('sidebarSearchInput');
        sidebarSearchInput.addEventListener('input', function() {{
            const query = this.value.toLowerCase().trim();
            const items = document.querySelectorAll('.author-item');
            
            items.forEach(item => {{
                const authorName = item.getAttribute('data-name');
                if (authorName.includes(query)) {{
                    item.style.display = 'flex';
                }} else {{
                    item.style.display = 'none';
                }}
            }});
        }});

        function copyCitation(btn, citationText) {{
            navigator.clipboard.writeText(citationText).then(() => {{
                const toast = document.getElementById('copyToast');
                toast.classList.add('show');
                
                const origHtml = btn.innerHTML;
                btn.innerHTML = '<i class="fa-solid fa-check"></i> {t["copied"]}';
                btn.style.backgroundColor = '#10b981';
                btn.style.color = 'white';
                btn.style.borderColor = '#10b981';
                
                setTimeout(() => {{
                    toast.classList.remove('show');
                }}, 2500);

                setTimeout(() => {{
                    btn.innerHTML = origHtml;
                    btn.style.backgroundColor = '';
                    btn.style.color = '';
                    btn.style.borderColor = '';
                }}, 2000);
            }}).catch(err => {{
                console.error('Failed to copy: ', err);
            }});
        }}

        window.addEventListener('DOMContentLoaded', () => {{
            const urlParams = new URLSearchParams(window.location.search);
            const authorParam = urlParams.get('author');
            if (authorParam) {{
                const items = document.querySelectorAll('.author-item');
                let foundBtn = null;
                items.forEach(item => {{
                    const authorName = item.querySelector('.author-name').innerText.trim();
                    if (authorName === authorParam) {{
                        foundBtn = item;
                    }}
                }});
                
                if (foundBtn) {{
                    selectAuthor(authorParam, foundBtn);
                    foundBtn.scrollIntoView({{ behavior: 'smooth', block: 'nearest' }});
                }}
            }}
        }});
    </script>
</body>
</html>
"""
    with open(output_path, "w", encoding="utf-8") as f:
        f.write(html_content)
    print(f"[{lang.upper()}] HTML Citations generated successfully.")

def run_export_for_lang(lang):
    json_path = f"outputs/faculty_directory_{lang}.json"
    excel_path = f"outputs/faculty_directory_{lang}.xlsx"
    dir_html_path = f"outputs/faculty_directory_{lang}.html"
    cit_html_path = f"outputs/faculty_citations_{lang}.html"
    
    if not os.path.exists(json_path):
        print(f"[{lang.upper()}] Cache file '{json_path}' not found, skipping.")
        return
        
    with open(json_path, "r", encoding="utf-8") as f:
        data = json.load(f)
        
    generate_excel(data, excel_path, lang)
    generate_html_directory(data, dir_html_path, lang)
    generate_html_citations(data, cit_html_path, lang)

def main():
    parser = argparse.ArgumentParser(description="Export scraped faculty data v2 to styled Excel & HTML")
    parser.add_argument("--lang", default="both", choices=["en", "tr", "both"], help="Language version to export")
    args = parser.parse_args()

    

    if args.lang in ["en", "both"]:
        run_export_for_lang("en")
        
    if args.lang in ["tr", "both"]:
        run_export_for_lang("tr")

    print("\nAll exports completed successfully!")

if __name__ == "__main__":
    main()
